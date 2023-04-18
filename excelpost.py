from lib2to3.pgen2 import driver
import openpyxl as OP
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait as WDW
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select 



excel_file = (r"C:\Users\mtl91475\Desktop\test.xlsx")

wb = OP.load_workbook(excel_file)
sheet = wb.active

sheet.cell(row=6,column=4).value="22XPB0781"
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(row=r,column=c).value="welcome"

wb.save(excel_file)

print(sheet.values)





